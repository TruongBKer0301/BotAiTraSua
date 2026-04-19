import logging
import asyncio
import os
import hashlib
import hmac
import re
import time
import unicodedata
from asyncio import to_thread
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional
from urllib.parse import quote_plus

import httpx
from dotenv import load_dotenv
from groq import Groq
from openpyxl import load_workbook
from telegram import Update
from telegram.error import Conflict, NetworkError
from telegram.ext import (
	Application,
	CommandHandler,
	ContextTypes,
	ConversationHandler,
	MessageHandler,
	filters,
)


logging.basicConfig(
	format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
	level=logging.INFO,
)
logger = logging.getLogger(__name__)

# Giảm spam log khi mạng chập chờn liên tục trong lúc long polling.
NETWORK_ERROR_LOG_COOLDOWN_SECONDS = 30
_last_network_error_log_ts = 0.0
PAYOS_STATUS_POLL_INTERVAL_SECONDS = int(os.getenv("PAYOS_STATUS_POLL_INTERVAL_SECONDS", "5"))
PAYOS_STATUS_POLL_TIMEOUT_SECONDS = int(os.getenv("PAYOS_STATUS_POLL_TIMEOUT_SECONDS", "900"))


MENU_XLSX_FILE = Path(__file__).with_name("Menu.xlsx")
DOTENV_FILE = Path(__file__).with_name(".env")
MENU_IMAGE_FILE = Path(__file__).with_name("Menu.png")
PAYOS_API_BASE_URL = os.getenv("PAYOS_API_BASE_URL", "https://api-merchant.payos.vn").rstrip("/")
PAYOS_RETURN_URL = os.getenv("PAYOS_RETURN_URL", "https://example.com/payos-return")
PAYOS_CANCEL_URL = os.getenv("PAYOS_CANCEL_URL", "https://example.com/payos-cancel")

load_dotenv(dotenv_path=DOTENV_FILE)

SELECT_CATEGORY, SELECT_SIZE, SELECT_QUANTITY, ADD_TOPPING, CONFIRM, PAYMENT_PHONE = range(6)
# Conversation states:
# SELECT_CATEGORY -> chọn món
# SELECT_SIZE -> chọn size
# ADD_TOPPING -> chọn topping
# SELECT_QUANTITY -> nhập số lượng (chỉ khi có topping)
# CONFIRM -> thêm món / xác nhận / hủy
# PAYMENT_PHONE -> nhập số điện thoại thanh toán


@dataclass
class MenuItem:
	category: str
	item_id: str
	name: str
	description: str
	price_m: int
	price_l: int
	available: bool


def parse_bool(value: str) -> bool:
	return str(value).strip().lower() in {"true", "1", "yes", "y"}


def _safe_cell_text(value: object) -> str:
	if value is None:
		return ""
	if isinstance(value, float) and value != value:
		return ""
	text = str(value).strip()
	return "" if text.lower() == "none" else text


def _menu_item_from_row(row: Dict[str, object]) -> MenuItem:
	category_value = (
		row.get("category")
		or row.get("Category")
		or row.get("Column1")
		or ""
	)

	return MenuItem(
		category=_safe_cell_text(category_value),
		item_id=_safe_cell_text(row.get("item_id", "")),
		name=_safe_cell_text(row.get("name", "")),
		description=_safe_cell_text(row.get("description", "")),
		price_m=int(float(_safe_cell_text(row.get("price_m", "0")))),
		price_l=int(float(_safe_cell_text(row.get("price_l", "0")))),
		available=parse_bool(_safe_cell_text(row.get("available", "true"))),
	)


def load_menu_from_xlsx(xlsx_file: Path) -> List[MenuItem]:
	# Đọc toàn bộ menu từ file Excel và bỏ qua dòng lỗi để bot vẫn khởi động được.
	menu_items: List[MenuItem] = []
	workbook = load_workbook(xlsx_file, data_only=True, read_only=True)
	sheet = workbook.active
	rows_iter = sheet.iter_rows(values_only=True)
	headers = next(rows_iter, None)
	if not headers:
		return menu_items

	header_names = [str(header).strip() if header is not None else "" for header in headers]
	rows = []
	for row in rows_iter:
		row_dict: Dict[str, object] = {}
		for index, header_name in enumerate(header_names):
			if not header_name:
				continue
			cell_value = row[index] if index < len(row) else None
			row_dict[header_name] = cell_value
			row_dict[header_name.lower()] = cell_value
		rows.append(row_dict)

	for row in rows:
		try:
			menu_items.append(_menu_item_from_row(row))
		except (KeyError, ValueError) as exc:
			logger.warning("Bo qua dong menu loi %s: %s", row, exc)

	logger.info("Doc menu XLSX thanh cong: %s", xlsx_file.name)
	return menu_items


def load_menu() -> List[MenuItem]:
	if MENU_XLSX_FILE.exists():
		return load_menu_from_xlsx(MENU_XLSX_FILE)

	raise FileNotFoundError(
		f"Khong tim thay file menu. Can co {MENU_XLSX_FILE.name}"
	)


MENU_ITEMS = load_menu()
ITEM_BY_ID: Dict[str, MenuItem] = {i.item_id: i for i in MENU_ITEMS if i.available}
TOPPINGS = [i for i in MENU_ITEMS if i.available and i.category == "Topping"]
DRINK_ITEMS = [i for i in MENU_ITEMS if i.available and i.category != "Topping"]
AI_GROQ_HINT = "Gợi ý nhanh: dùng /goiy <tất cả yêu cầu của bạn>"


def with_ai_hint(text: str) -> str:
	content = (text or "").strip()
	if not content:
		return AI_GROQ_HINT
	if AI_GROQ_HINT in content:
		return content
	return f"{content}\n\n{AI_GROQ_HINT}"


async def reply_text_with_hint(message, text: str, **kwargs) -> None:
	await message.reply_text(with_ai_hint(text), **kwargs)


async def send_message_with_hint(chat, text: str, **kwargs) -> None:
	await chat.send_message(with_ai_hint(text), **kwargs)


async def send_bot_message_with_hint(bot, chat_id: int, text: str, **kwargs) -> None:
	await bot.send_message(chat_id=chat_id, text=with_ai_hint(text), **kwargs)


async def reply_photo_with_hint(message, photo, caption: str = "", **kwargs) -> None:
	await message.reply_photo(photo=photo, caption=with_ai_hint(caption), **kwargs)


def get_groq_client() -> Optional[Groq]:
	api_key = os.getenv("GROQ_API_KEY", "").strip()
	if not api_key:
		return None
	return Groq(api_key=api_key)


def build_menu_context() -> str:
	# Chuỗi menu đầy đủ dùng cho prompt AI gợi ý.
	drink_lines = []
	for item in MENU_ITEMS:
		if not item.available:
			continue
		if item.category == "Topping":
			continue
		drink_lines.append(
			f"- {item.name} ({item.item_id}) | {item.category} | M:{item.price_m} VND | L:{item.price_l} VND | {item.description}"
		)

	topping_lines = []
	for top in TOPPINGS:
		topping_lines.append(
			f"- {top.name} ({top.item_id}) | {top.price_m} VND"
		)

	menu_text = "\n".join(drink_lines)
	topping_text = "\n".join(topping_lines)
	return (
		"MENU DO UONG:\n"
		f"{menu_text}\n\n"
		"MENU TOPPING:\n"
		f"{topping_text}"
	)


def build_menu_preview() -> str:
	lines: List[str] = ["Menu hôm nay:"]
	categories = sorted({i.category for i in DRINK_ITEMS})
	for category in categories:
		lines.append(f"\n{category}:")
		for item in DRINK_ITEMS:
			if item.category != category:
				continue
			lines.append(
				f"- {item.name} ({item.item_id}) | M {format_price(item.price_m)} | L {format_price(item.price_l)}"
			)

	if TOPPINGS:
		lines.append("\nTopping:")
		for top in TOPPINGS:
			lines.append(f"- {top.name} ({top.item_id}) | {format_price(top.price_m)}")

	lines.append("\nBạn có thể nhập số, tên món hoặc mã món để chọn, sau đó chọn size và topping.")
	return "\n".join(lines)


async def send_menu_then_options(chat) -> None:
	# /start: gửi ảnh menu trước, sau đó gửi hướng dẫn đặt món.
	if MENU_IMAGE_FILE.exists():
		with open(MENU_IMAGE_FILE, "rb") as photo:
			await chat.send_photo(photo=photo)
	else:
		await send_message_with_hint(chat, build_menu_preview())

	await send_message_with_hint(chat, 
		"VUI LÒNG CHỌN MÓN ❤️.\n"
		"Bạn có thể nhập gộp tên món, size và topping trong một tin nhắn.\n"
		"Ví dụ: 5 hồng trà trân châu 3 size m 2 size l, 2 kem tươi, 2 sữa dừa\n"
		"Nếu muốn hủy, hãy gõ /cancel",
	)


async def send_menu_image(chat) -> None:
	# Mọi yêu cầu "xin menu" đều đi qua hàm này.
	if MENU_IMAGE_FILE.exists():
		with open(MENU_IMAGE_FILE, "rb") as photo:
			await chat.send_photo(photo=photo)
		await send_message_with_hint(chat, "Đã gửi menu cho bạn.")
		return

	await send_message_with_hint(chat, build_menu_preview())


def wants_menu_image(text: str) -> bool:
	normalized = normalize_text(text)
	return any(
		phrase in normalized
		for phrase in (
			"xin menu",
			"cho xin menu",
			"xem lai menu",
			"xem menu",
			"cho xem menu",
			"gui menu",
			"menu",
		)
	)


def wants_choose_item_restart(text: str) -> bool:
	normalized = normalize_text(text)
	return any(
		phrase in normalized
		for phrase in (
			"toi muon chon mon",
			"chon mon",
			"doi mon",
			"mon nao",
			"mon nao ban",
			"chon lai mon",
		)
	)


def wants_add_topping_request(text: str) -> bool:
	normalized = normalize_text(text)
	return any(
		phrase in normalized
		for phrase in (
			"them topping",
			"them toping",
			"them tp",
			"them tran chau",
			"topping",
			"toping",
		)
	)


def get_payos_config() -> tuple[Optional[str], Optional[str], Optional[str]]:
	client_id = os.getenv("PAYOS_CLIENT_ID", "").strip()
	api_key = os.getenv("PAYOS_API_KEY", "").strip()
	checksum_key = os.getenv("PAYOS_CHECKSUM_KEY", "").strip()
	if not client_id or not api_key or not checksum_key:
		return None, None, None
	return client_id, api_key, checksum_key


def sanitize_phone_number(text: str) -> Optional[str]:
	digits = re.sub(r"\D+", "", text or "")
	if len(digits) < 9:
		return None
	return digits


def build_cart_total(context: ContextTypes.DEFAULT_TYPE) -> int:
	cart: List[Dict[str, object]] = context.user_data.get("cart", [])
	total = 0
	for row in cart:
		item = ITEM_BY_ID.get(str(row.get("item_id", "")))
		if not item:
			continue
		size = str(row.get("size", "M"))
		quantity = int(row.get("quantity", 1))
		topping_ids = [str(t) for t in row.get("toppings", [])]
		total += calc_order_total(item, size, quantity, topping_ids)
	return total


def build_payos_items(context: ContextTypes.DEFAULT_TYPE) -> List[Dict[str, object]]:
	cart: List[Dict[str, object]] = context.user_data.get("cart", [])
	items: List[Dict[str, object]] = []
	for row in cart:
		item = ITEM_BY_ID.get(str(row.get("item_id", "")))
		if not item:
			continue
		size = str(row.get("size", "M"))
		quantity = int(row.get("quantity", 1))
		topping_ids = [str(t) for t in row.get("toppings", [])]
		base_price = item.price_m if size == "M" else item.price_l
		topping_total = sum(ITEM_BY_ID[t].price_m for t in topping_ids if t in ITEM_BY_ID)
		items.append(
			{
				"name": f"{item.name} size {size}",
				"quantity": quantity,
				"price": base_price + topping_total,
				"unit": "ly",
			}
		)
	return items


def build_payos_signature(amount: int, cancel_url: str, description: str, order_code: int, return_url: str, checksum_key: str) -> str:
	data = (
		f"amount={amount}&"
		f"cancelUrl={cancel_url}&"
		f"description={description}&"
		f"orderCode={order_code}&"
		f"returnUrl={return_url}"
	)
	return hmac.new(checksum_key.encode("utf-8"), data.encode("utf-8"), hashlib.sha256).hexdigest()


def create_payos_payment_link(context: ContextTypes.DEFAULT_TYPE, phone_number: str) -> Dict[str, object]:
	# Tạo payment link PayOS dựa trên giỏ hàng hiện tại + số điện thoại user nhập.
	client_id, api_key, checksum_key = get_payos_config()
	if not client_id or not api_key or not checksum_key:
		raise RuntimeError("Thiếu PAYOS_CLIENT_ID, PAYOS_API_KEY hoặc PAYOS_CHECKSUM_KEY trong .env")

	amount = build_cart_total(context)
	order_code = int(time.time() * 1000)
	description = phone_number
	signature = build_payos_signature(amount, PAYOS_CANCEL_URL, description, order_code, PAYOS_RETURN_URL, checksum_key)
	payload = {
		"orderCode": order_code,
		"amount": amount,
		"description": description,
		"buyerPhone": phone_number,
		"items": build_payos_items(context),
		"cancelUrl": PAYOS_CANCEL_URL,
		"returnUrl": PAYOS_RETURN_URL,
		"signature": signature,
	}

	response = httpx.post(
		f"{PAYOS_API_BASE_URL}/v2/payment-requests",
		headers={
			"x-client-id": client_id,
			"x-api-key": api_key,
			"Content-Type": "application/json",
		},
		json=payload,
		timeout=30,
	)
	response.raise_for_status()
	data = response.json()
	if str(data.get("code", "")) != "00":
		raise RuntimeError(f"PayOS trả lỗi: {data.get('desc', 'unknown error')}")
	return data.get("data", {})


def build_qr_image_url(qr_data: str) -> str:
	return f"https://api.qrserver.com/v1/create-qr-code/?size=320x320&data={quote_plus(qr_data)}"


def get_payos_payment_status(order_code: str) -> str:
	client_id, api_key, _ = get_payos_config()
	if not client_id or not api_key:
		raise RuntimeError("Thiếu cấu hình PayOS để kiểm tra trạng thái thanh toán")

	response = httpx.get(
		f"{PAYOS_API_BASE_URL}/v2/payment-requests/{order_code}",
		headers={
			"x-client-id": client_id,
			"x-api-key": api_key,
			"Content-Type": "application/json",
		},
		timeout=30,
	)
	response.raise_for_status()
	payload = response.json()
	if str(payload.get("code", "")) != "00":
		raise RuntimeError(f"PayOS trả lỗi khi kiểm tra trạng thái: {payload.get('desc', 'unknown error')}")

	data = payload.get("data") or {}
	status_candidates = (
		data.get("status"),
		data.get("paymentStatus"),
		data.get("transactionStatus"),
	)
	for status in status_candidates:
		if status:
			return str(status).strip().upper()
	return ""


async def monitor_payos_payment(
	context: ContextTypes.DEFAULT_TYPE,
	chat_id: int,
	order_code: str,
) -> None:
	deadline = time.time() + max(PAYOS_STATUS_POLL_TIMEOUT_SECONDS, 60)
	paid_statuses = {"PAID", "SUCCESS", "SUCCEEDED", "COMPLETED"}
	stop_statuses = {"CANCELLED", "CANCELED", "EXPIRED", "FAILED"}

	while time.time() < deadline:
		try:
			status = await to_thread(get_payos_payment_status, order_code)
		except Exception as exc:
			logger.warning("Khong kiem tra duoc trang thai thanh toan cho don %s: %s", order_code, exc)
			await asyncio.sleep(max(PAYOS_STATUS_POLL_INTERVAL_SECONDS, 3))
			continue

		if status in paid_statuses:
			await send_bot_message_with_hint(context.bot, 
				chat_id=chat_id,
				text="Bạn đã chuyển khoản thành công, Quán sẽ gọi cho bạn ngay !",
			)
			return

		if status in stop_statuses:
			logger.info("Dung theo doi don %s do trang thai %s", order_code, status)
			return

		await asyncio.sleep(max(PAYOS_STATUS_POLL_INTERVAL_SECONDS, 3))

	logger.info("Het thoi gian theo doi thanh toan cho don %s", order_code)


def ask_groq_for_recommendation(user_request: str) -> str:
	client = get_groq_client()
	if client is None:
		return (
			"Chưa có GROQ_API_KEY.\n"
			"Hãy thêm vào file .env:\n"
			"GROQ_API_KEY=your_groq_api_key"
		)

	menu_context = build_menu_context()
	system_prompt = (
		"Bạn là nhân viên tư vấn đồ uống cho quán nước. "
		"Chỉ được gợi ý dựa trên menu cung cấp. "
		"Trả lời bằng tiếng Việt, ngắn gọn, dễ hiểu. "
		"Phải phân tích và bám đầy đủ tất cả yêu cầu, ràng buộc và thông tin mà khách nêu ra, "
		"không chỉ dựa vào sở thích. "
		"Mỗi gợi ý nên có tên món, size đề xuất, topping đề xuất và ước tính giá."
	)

	completion = client.chat.completions.create(
		model="llama-3.3-70b-versatile",
		temperature=0.5,
		messages=[
			{"role": "system", "content": system_prompt},
			{
				"role": "user",
				"content": (
					f"{menu_context}\n\n"
					f"Tất cả yêu cầu của khách: {user_request}\n"
					"Hãy đề xuất 3 lựa chọn phù hợp nhất, bám đủ mọi ràng buộc đã nêu."
				),
			},
		],
	)

	content = completion.choices[0].message.content
	if not content:
		return "Không nhận được phản hồi từ Groq. Bạn thử lại sau."
	return content.strip()


def ask_groq_for_order_prompt(stage: str, item: Optional[MenuItem] = None) -> str:
	def fallback_prompt() -> str:
		if stage == "size_topping" and item:
			return (
				f"{item.name} ({item.item_id})\n"
				f"- Size M: {format_price(item.price_m)}\n"
				f"- Size L: {format_price(item.price_l)}\n"
				"Nhập size và topping trong 1 tin nhắn. Ví dụ: M, topping trân châu"
			)
		if stage == "topping_followup":
			return "Mình cần topping cho món này. Nếu không thêm, hãy nhập 'không'."
		return "Bạn muốn làm gì tiếp theo?"

	client = get_groq_client()
	if client is None:
		return fallback_prompt()

	system_prompt = (
		"Bạn là trợ lý AI cho quán nước. "
		"Chỉ trả lời ngắn gọn, tự nhiên, lịch sự bằng tiếng Việt. "
		"Mục tiêu là hỏi khách đúng 1 thông tin còn thiếu trong luồng đặt món và bám đúng tất cả yêu cầu hiện có. "
		"Không giải thích dài dòng."
	)

	if stage == "size_topping" and item:
		user_prompt = (
			f"Khách vừa chọn món {item.name} ({item.item_id}). "
			f"Hãy yêu cầu khách nhập size và topping luôn trong một tin nhắn. "
			"Nếu chưa có topping thì gợi ý hỏi khách có cần topping không."
		)
	elif stage == "topping_followup":
		user_prompt = "Hãy hỏi khách có cần topping không. Nếu có, hãy bảo họ nhập tên topping hoặc mã topping; nếu không thì nhập 'không'."
	else:
		user_prompt = "Hãy hỏi khách một câu ngắn để tiếp tục đặt món."

	try:
		completion = client.chat.completions.create(
			model="llama-3.3-70b-versatile",
			temperature=0.4,
			messages=[
				{"role": "system", "content": system_prompt},
				{"role": "user", "content": user_prompt},
			],
		)
	except Exception as exc:
		logger.warning("Groq prompt order bi loi, dung cau mac dinh: %s", exc)
		return fallback_prompt()

	content = completion.choices[0].message.content
	if not content:
		return fallback_prompt()
	return content.strip()


def is_order_in_progress(context: ContextTypes.DEFAULT_TYPE) -> bool:
	return any(
		key in context.user_data
		for key in (
			"selected_item_ids",
			"pending_item_ids",
			"current_item_id",
			"current_size",
			"current_quantity",
			"current_toppings",
			"cart",
		)
	)


def format_price(vnd: int) -> str:
	return f"{vnd:,.0f}d".replace(",", ".")

def parse_item_selection_numbers(text: str) -> List[str]:
	tokens = text.replace(",", " ").split()
	if not tokens:
		raise ValueError("Bạn chưa nhập số nào")

	selected_item_ids: List[str] = []
	seen = set()
	for token in tokens:
		if not token.isdigit():
			raise ValueError("Chỉ được nhập số, cách nhau bằng dấu cách")

		index = int(token)
		if index < 1 or index > len(DRINK_ITEMS):
			raise ValueError(f"Số {index} không hợp lệ. Chọn từ 1 đến {len(DRINK_ITEMS)}")

		item_id = DRINK_ITEMS[index - 1].item_id
		if item_id not in seen:
			seen.add(item_id)
			selected_item_ids.append(item_id)

	return selected_item_ids


def parse_item_selection_input(text: str) -> List[str]:
	# Chấp nhận 2 kiểu nhập:
	# 1) Danh sách số thứ tự (ví dụ: "1 3 5")
	# 2) Câu tự nhiên chứa tên món/mã món (ví dụ: "2 cà phê mocha size m")
	compact = (text or "").strip()
	if not compact:
		raise ValueError("Bạn chưa nhập món nào")

	tokens = compact.replace(",", " ").split()
	if tokens and all(token.isdigit() for token in tokens):
		return parse_item_selection_numbers(compact)

	normalized = apply_item_aliases(normalize_text(compact))
	tokens = [token for token in re.split(r"[^a-z0-9]+", normalized) if token]
	selected_item_ids: List[str] = []
	for item in sorted(DRINK_ITEMS, key=lambda entry: len(normalize_text(entry.name)), reverse=True):
		# Sort tên dài trước để giảm match nhầm khi tên món có phần giống nhau.
		item_name = apply_item_aliases(normalize_text(item.name))
		name_pattern = rf"(?:^|\s){re.escape(item_name)}(?:\s|$)"
		if (item_name and re.search(name_pattern, normalized)) or item.item_id.lower() in tokens:
			selected_item_ids.append(item.item_id)

	selected_item_ids = dedupe_preserve_order(selected_item_ids)
	if not selected_item_ids:
		raise ValueError("Chưa nhận diện được món. Hãy nhập số, tên món hoặc mã món.")

	return selected_item_ids


def apply_item_aliases(normalized_text: str) -> str:
	# Gom các biến thể người dùng hay gõ: caphe/cf/coffee -> cà phê.
	text = f" {normalized_text} "
	text = re.sub(r"\bcaphe\b", "ca phe", text)
	text = re.sub(r"\bcoffee\b", "ca phe", text)
	text = re.sub(r"\bcf\b", "ca phe", text)
	text = re.sub(r"\bcafe\b", "ca phe", text)
	return re.sub(r"\s+", " ", text).strip()


def is_no_topping_text(normalized_text: str) -> bool:
	return normalized_text in {
		"khong",
		"ko",
		"no",
		"none",
		"khong co",
		"khong topping",
		"khong co topping",
		"khoong",
	}


def get_progress_text(context: ContextTypes.DEFAULT_TYPE) -> str:
	cart_count = len(context.user_data.get("cart", []))
	pending_count = len(context.user_data.get("pending_item_ids", []))
	total = cart_count + pending_count
	current = cart_count + 1
	return f"Món {current}/{total}" if total > 0 else "Món hiện tại"


def build_size_topping_prompt(item: MenuItem) -> str:
	return (
		f"Tên món: {item.name} ({item.item_id})\n"
		f"Size M: {format_price(item.price_m)} | Size L: {format_price(item.price_l)}\n"
		"Nhập size và topping trong 1 tin nhắn. Ví dụ: M, topping trân châu"
	)


def normalize_text(text: str) -> str:
	text = unicodedata.normalize("NFD", text.lower())
	text = text.replace("đ", "d")
	text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
	return re.sub(r"\s+", " ", text).strip()


def dedupe_preserve_order(items: List[str]) -> List[str]:
	seen = set()
	result: List[str] = []
	for item in items:
		if item not in seen:
			seen.add(item)
			result.append(item)
	return result


def find_topping_ids(normalized: str, tokens: List[str]) -> List[str]:
	topping_ids: List[str] = []
	for top in TOPPINGS:
		top_name = normalize_text(top.name)
		if (top_name and top_name in normalized) or top.item_id.lower() in tokens:
			topping_ids.append(top.item_id)
	return dedupe_preserve_order(topping_ids)


def parse_size_and_topping_input(text: str) -> tuple[str, List[str]]:
	# Đọc size từ cụm "size m/l" hoặc token ngắn gọn "m/l", "nhỏ/lớn".
	normalized = normalize_text(text)
	tokens = [token for token in re.split(r"[^a-z0-9]+", normalized) if token]
	size: Optional[str] = None
	for index, token in enumerate(tokens):
		if token == "size" and index + 1 < len(tokens):
			next_token = tokens[index + 1]
			if next_token == "m":
				size = "M"
				break
			if next_token == "l":
				size = "L"
				break
		if token == "m":
			size = "M"
			break
		if token in {"nho", "be"}:
			size = "M"
			break
		if token == "l":
			size = "L"
			break
		if token in {"lon", "to"}:
			size = "L"
			break

	if size is None:
		raise ValueError("Thiếu size. Vui lòng nhập M hoặc L.")

	return size, find_topping_ids(normalized, tokens)


def parse_topping_input(text: str) -> List[str]:
	normalized = normalize_text(text)
	tokens = [token for token in re.split(r"[^a-z0-9]+", normalized) if token]
	if is_no_topping_text(normalized):
		return []

	topping_ids = extract_topping_ids_with_quantity(text)
	if not topping_ids:
		topping_ids = find_topping_ids(normalized, tokens)
	if not topping_ids:
		raise ValueError("Chưa nhận diện được topping. Vui lòng nhập tên topping hoặc gõ 'không'.")

	return topping_ids


def extract_size_quantity_variants(text: str) -> List[tuple[str, int]]:
	# Hỗ trợ câu gồm nhiều biến thể size trong 1 món, ví dụ: "3 size m 2 size l".
	normalized = normalize_text(text)
	matches = re.findall(r"(\d+)\s*size\s*([ml])", normalized)
	variants: List[tuple[str, int]] = []
	for qty_text, size_text in matches:
		qty = int(qty_text)
		if qty > 0:
			variants.append((size_text.upper(), qty))
	return variants


def extract_leading_quantity(text: str) -> Optional[int]:
	normalized = normalize_text(text)
	match = re.match(r"^(\d+)\b", normalized)
	if not match:
		return None
	qty = int(match.group(1))
	return qty if qty > 0 else None


def extract_topping_ids_with_quantity(text: str) -> List[str]:
	# Topping cho phép nhập kèm số lượng, ví dụ: "2 kem tươi, 1 thạch xanh".
	normalized = normalize_text(text)
	if is_no_topping_text(normalized):
		return []

	topping_ids: List[str] = []
	for top in TOPPINGS:
		top_name = normalize_text(top.name)
		exact_pattern = rf"(\d+)\s*(?:topping\s*)?{re.escape(top_name)}"
		qty_matches = re.findall(exact_pattern, normalized)
		total_qty = sum(int(q) for q in qty_matches)

		if total_qty == 0 and top_name in normalized:
			total_qty = 1

		for _ in range(total_qty):
			topping_ids.append(top.item_id)

	return topping_ids


def build_cart_summary(context: ContextTypes.DEFAULT_TYPE) -> str:
	# Tạo hóa đơn dạng gạch đầu dòng đơn giản.
	cart: List[Dict[str, object]] = context.user_data.get("cart", [])
	if not cart:
		return "Giỏ hàng đang trống."

	lines = []
	grand_total = 0

	for row in cart:
		item = ITEM_BY_ID.get(str(row["item_id"]))
		if not item:
			continue

		size = str(row["size"])
		quantity = int(row["quantity"])
		topping_ids = [str(t) for t in row.get("toppings", [])]
		line_total = calc_order_total(item, size, quantity, topping_ids)
		grand_total += line_total

		topping_text = "không"
		if topping_ids:
			topping_count: Dict[str, int] = {}
			for topping_id in topping_ids:
				topping_count[topping_id] = topping_count.get(topping_id, 0) + 1

			topping_parts: List[str] = []
			for topping_id, count in topping_count.items():
				top_item = ITEM_BY_ID.get(topping_id)
				if not top_item:
					continue
				if count > 1:
					topping_parts.append(f"{top_item.name.lower()} x{count}")
				else:
					topping_parts.append(top_item.name.lower())
			if topping_parts:
				topping_text = ", ".join(topping_parts)

		lines.append(
			f"- {item.name.lower()} - Size : {size.upper()} - SL :{quantity} - Topping: {topping_text} - Thành tiền : {format_price(line_total)}"
		)

	lines.append(f"- Tổng cộng - Thành tiền : {format_price(grand_total)}")
	return "\n".join(lines)


def format_topping_ids_for_text(topping_ids: List[str]) -> str:
	topping_count: Dict[str, int] = {}
	for topping_id in topping_ids:
		topping_count[topping_id] = topping_count.get(topping_id, 0) + 1

	parts: List[str] = []
	for topping_id, count in topping_count.items():
		top_item = ITEM_BY_ID.get(topping_id)
		if not top_item:
			continue
		if count > 1:
			parts.append(f"{top_item.name.lower()} x{count}")
		else:
			parts.append(top_item.name.lower())

	return ", ".join(parts) if parts else "không"


def add_current_item_to_cart(context: ContextTypes.DEFAULT_TYPE, topping_ids: List[str]) -> None:
	# Chot 1 dong gio hang roi xoa cac bien tam cua mon hien tai.
	item_id = context.user_data.get("current_item_id")
	size = context.user_data.get("current_size")
	quantity = context.user_data.get("current_quantity")
	if not item_id or not size or not quantity:
		return

	cart: List[Dict[str, object]] = context.user_data.setdefault("cart", [])
	cart.append(
		{
			"item_id": item_id,
			"size": size,
			"quantity": int(quantity),
			"toppings": topping_ids.copy(),
		}
	)

	pending_ids: List[str] = context.user_data.get("pending_item_ids", [])
	if pending_ids and pending_ids[0] == item_id:
		pending_ids.pop(0)
	elif item_id in pending_ids:
		pending_ids.remove(item_id)

	context.user_data["current_item_id"] = None
	context.user_data["current_size"] = None
	context.user_data["current_quantity"] = None
	context.user_data["current_toppings"] = []


def calc_order_total(item: MenuItem, size: str, quantity: int, topping_ids: List[str]) -> int:
	base_price = item.price_m if size == "M" else item.price_l
	topping_total = sum(ITEM_BY_ID[t].price_m for t in topping_ids if t in ITEM_BY_ID)
	return (base_price + topping_total) * quantity


def build_order_summary(context: ContextTypes.DEFAULT_TYPE) -> str:
	return build_cart_summary(context)


async def prompt_next_item_or_confirm(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
	# Nếu còn món đang chờ -> hỏi size/topping món tiếp theo.
	# Nếu hết món -> hiện tổng đơn và chờ xác nhận.
	pending_item_ids = context.user_data.get("pending_item_ids", [])
	if pending_item_ids:
		next_item = ITEM_BY_ID.get(pending_item_ids[0])
		if not next_item:
			if update.message:
				await reply_text_with_hint(update.message, "Có lỗi dữ liệu món. Gõ /start để đặt lại.")
			return ConversationHandler.END

		context.user_data["current_item_id"] = next_item.item_id
		context.user_data["awaiting_size_selection"] = True
		context.user_data["awaiting_topping_selection"] = False
		progress = get_progress_text(context)
		if update.message:
			await reply_text_with_hint(update.message, 
				f"{progress}: {next_item.name}\n"
				f"Mô tả: {next_item.description}\n"
				f"{build_size_topping_prompt(next_item)}\n"
				f"{ask_groq_for_order_prompt('size_topping', next_item)}\n"
				"(Gõ /cancel để hủy)",
			)
		return SELECT_SIZE

	summary = build_order_summary(context)
	# Về bước xác nhận thì dọn cờ tạm để tránh đi nhầm luồng size/topping.
	context.user_data["awaiting_size_selection"] = False
	context.user_data["awaiting_topping_selection"] = False
	context.user_data.pop("awaiting_confirm_topping_name", None)
	context.user_data.pop("awaiting_confirm_topping_quantity", None)
	context.user_data.pop("pending_confirm_topping_ids", None)
	if update.message:
		await reply_text_with_hint(update.message, 
			f"{summary}\n\n"
			"Bạn có muốn đặt thêm món không?\n"
			"- Nhập tên/mã/số món để thêm món\n"
			"- Nhập 'xác nhận' để sang bước thanh toán\n"
			"- Nhập 'hủy' để hủy đơn"
		)
	return CONFIRM


async def start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
	context.user_data.clear()
	context.user_data["awaiting_item_selection"] = True
	if update.message:
		await send_menu_then_options(update.message.chat)
	return SELECT_CATEGORY


async def show_main_menu(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
	if update.message:
		await send_menu_image(update.message.chat)

	if not is_order_in_progress(context):
		context.user_data["awaiting_item_selection"] = True
		return SELECT_CATEGORY

	if context.user_data.get("awaiting_payment_phone"):
		return PAYMENT_PHONE
	if context.user_data.get("awaiting_size_selection"):
		return SELECT_SIZE
	if context.user_data.get("awaiting_topping_selection"):
		return ADD_TOPPING
	if context.user_data.get("current_item_id") and context.user_data.get("current_size"):
		return SELECT_QUANTITY
	if context.user_data.get("awaiting_item_selection"):
		return SELECT_CATEGORY
	if context.user_data.get("cart") and not context.user_data.get("pending_item_ids"):
		return CONFIRM
	return SELECT_CATEGORY


async def on_select_items_by_text(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
	if not update.message:
		return SELECT_CATEGORY

	text = (update.message.text or "").strip()
	if wants_menu_image(text):
		# Đứng ở bước chọn món mà xin menu thì chỉ gửi ảnh, không đổi state.
		await send_menu_image(update.message.chat)
		return SELECT_CATEGORY

	try:
		selected_item_ids = parse_item_selection_input(text)
	except ValueError as exc:
		await reply_text_with_hint(update.message, str(exc))
		await send_menu_image(update.message.chat)
		return SELECT_CATEGORY

	context.user_data["selected_item_ids"] = selected_item_ids.copy()
	context.user_data["pending_item_ids"] = selected_item_ids.copy()
	context.user_data["cart"] = []
	context.user_data["awaiting_item_selection"] = False

	item = ITEM_BY_ID.get(selected_item_ids[0])
	if item is None:
		await reply_text_with_hint(update.message, "Món không hợp lệ. Gõ /start để đặt lại.")
		return ConversationHandler.END

	context.user_data["current_item_id"] = item.item_id
	context.user_data["awaiting_size_selection"] = True
	context.user_data["awaiting_topping_selection"] = False

	if len(selected_item_ids) == 1 and extract_size_quantity_variants(text):
		# Hỗ trợ câu full-order ngay sau menu: tên món + số lượng size + topping.
		return await on_select_size_by_text(update, context)

	if len(selected_item_ids) == 1:
		leading_qty = extract_leading_quantity(text)
		if leading_qty:
			try:
				size, topping_ids = parse_size_and_topping_input(text)
			except ValueError:
				pass
			else:
				context.user_data["current_size"] = size
				context.user_data["current_quantity"] = leading_qty
				context.user_data["current_toppings"] = topping_ids
				context.user_data["awaiting_size_selection"] = False

				if topping_ids:
					context.user_data["awaiting_topping_selection"] = False
					add_current_item_to_cart(context, topping_ids)
					return await prompt_next_item_or_confirm(update, context)

				add_current_item_to_cart(context, [])
				return await prompt_next_item_or_confirm(update, context)

	progress = get_progress_text(context)
	await reply_text_with_hint(update.message, 
		f"{progress}: {item.name}\n"
		f"Mô tả: {item.description}\n"
		f"{build_size_topping_prompt(item)}\n"
		f"{ask_groq_for_order_prompt('size_topping', item)}\n"
		"(Gõ /cancel để hủy)",
	)
	return SELECT_SIZE


async def on_select_size_by_text(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
	if not update.message:
		return SELECT_SIZE

	text = (update.message.text or "").strip()
	if wants_menu_image(text):
		await send_menu_image(update.message.chat)
		return SELECT_SIZE

	size_variants = extract_size_quantity_variants(text)
	if size_variants:
		current_item_id = context.user_data.get("current_item_id")
		if not current_item_id or current_item_id not in ITEM_BY_ID:
			await reply_text_with_hint(update.message, "Không xác định được món hiện tại. Gõ /start để đặt lại.")
			return ConversationHandler.END

		topping_ids = extract_topping_ids_with_quantity(text)
		if not topping_ids:
			# Câu full-order có sẵn số lượng theo size mà không nêu topping -> mặc định không topping.
			for size, qty in size_variants:
				context.user_data["current_item_id"] = current_item_id
				context.user_data["current_size"] = size
				context.user_data["current_quantity"] = qty
				context.user_data["current_toppings"] = []
				add_current_item_to_cart(context, [])
			return await prompt_next_item_or_confirm(update, context)

		for size, qty in size_variants:
			context.user_data["current_item_id"] = current_item_id
			context.user_data["current_size"] = size
			context.user_data["current_quantity"] = qty
			context.user_data["current_toppings"] = topping_ids
			add_current_item_to_cart(context, topping_ids)
		return await prompt_next_item_or_confirm(update, context)

	try:
		size, topping_ids = parse_size_and_topping_input(text)
	except ValueError as exc:
		current_item_id = context.user_data.get("current_item_id")
		item = ITEM_BY_ID.get(str(current_item_id)) if current_item_id else None
		if item:
			await reply_text_with_hint(update.message, 
				f"{exc}\n{build_size_topping_prompt(item)}\n{ask_groq_for_order_prompt('size_topping', item)}"
			)
		else:
			await reply_text_with_hint(update.message, str(exc))
		return SELECT_SIZE

	context.user_data["current_size"] = size
	context.user_data["awaiting_size_selection"] = False
	context.user_data["current_toppings"] = topping_ids
	if topping_ids:
		context.user_data["awaiting_topping_selection"] = False
		await reply_text_with_hint(update.message, "Anh/chị muốn đặt topping này với số lượng bao nhiêu ạ?")
		return SELECT_QUANTITY

	context.user_data["awaiting_topping_selection"] = True
	await reply_text_with_hint(update.message, ask_groq_for_order_prompt("topping_followup"))
	return ADD_TOPPING


async def on_quantity_input(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
	text = (update.message.text or "").strip()
	if wants_menu_image(text):
		await send_menu_image(update.message.chat)
		return SELECT_QUANTITY

	if wants_choose_item_restart(text):
		context.user_data.clear()
		context.user_data["awaiting_item_selection"] = True
		await send_menu_then_options(update.message.chat)
		return SELECT_CATEGORY

	# Trường hợp thêm topping ở bước CONFIRM: hỏi số lượng rồi mới cộng topping.
	if context.user_data.get("awaiting_confirm_topping_quantity"):
		match = re.search(r"\b\d+\b", normalize_text(text))
		if not match:
			try:
				updated_topping_ids = parse_topping_input(text)
			except ValueError:
				updated_topping_ids = []

			if updated_topping_ids:
				context.user_data["pending_confirm_topping_ids"] = updated_topping_ids
				await reply_text_with_hint(update.message, 
					f"Đã nhận topping: {format_topping_ids_for_text(updated_topping_ids)}. "
					"Vui lòng nhập số lượng topping."
				)
				return SELECT_QUANTITY

			await reply_text_with_hint(update.message, "Số lượng topping không hợp lệ. Vui lòng nhập số nguyên dương.")
			return SELECT_QUANTITY

		quantity = int(match.group(0))
		if quantity <= 0:
			await reply_text_with_hint(update.message, "Số lượng topping không hợp lệ. Vui lòng nhập số nguyên dương.")
			return SELECT_QUANTITY

		cart: List[Dict[str, object]] = context.user_data.get("cart", [])
		confirm_topping_ids: List[str] = context.user_data.get("pending_confirm_topping_ids", [])
		if not cart or not confirm_topping_ids:
			context.user_data.pop("awaiting_confirm_topping_quantity", None)
			context.user_data.pop("pending_confirm_topping_ids", None)
			await reply_text_with_hint(update.message, "Không tìm thấy món để thêm topping. Vui lòng chọn món lại.")
			return CONFIRM

		last_row = cart[-1]
		existing_toppings = [str(t) for t in last_row.get("toppings", [])]
		for topping_id in confirm_topping_ids:
			existing_toppings.extend([topping_id] * quantity)
		last_row["toppings"] = existing_toppings

		context.user_data.pop("awaiting_confirm_topping_quantity", None)
		context.user_data.pop("pending_confirm_topping_ids", None)

		summary = build_order_summary(context)
		await reply_text_with_hint(update.message, 
			f"Đã thêm topping cho món gần nhất.\n\n{summary}\n\n"
			"Bạn có muốn đặt thêm món không?\n"
			"- Nhập tên/mã/số món để thêm món\n"
			"- Nhập 'xác nhận' để sang bước thanh toán\n"
			"- Nhập 'hủy' để hủy đơn"
		)
		return CONFIRM

	if not text.isdigit() or int(text) <= 0:
		match = re.search(r"\b\d+\b", normalize_text(text))
		if match:
			quantity = int(match.group(0))
			if quantity > 0:
				current_item_id = context.user_data.get("current_item_id")
				if not current_item_id or current_item_id not in ITEM_BY_ID:
					if context.user_data.get("cart") and not context.user_data.get("pending_item_ids"):
						# Tránh rơi lỗi khi state còn SELECT_QUANTITY nhưng thực tế đã quay về CONFIRM.
						return await prompt_next_item_or_confirm(update, context)
					await reply_text_with_hint(update.message, "Không xác định được món hiện tại. Gõ /start để đặt lại.")
					return ConversationHandler.END

				context.user_data["current_quantity"] = quantity
				add_current_item_to_cart(context, context.user_data.get("current_toppings", []))
				return await prompt_next_item_or_confirm(update, context)

		await reply_text_with_hint(update.message, "Số lượng không hợp lệ. Vui lòng nhập số nguyên dương.")
		return SELECT_QUANTITY

	current_item_id = context.user_data.get("current_item_id")
	if not current_item_id or current_item_id not in ITEM_BY_ID:
		if context.user_data.get("cart") and not context.user_data.get("pending_item_ids"):
			# Tránh rơi lỗi khi state còn SELECT_QUANTITY nhưng thực tế đã quay về CONFIRM.
			return await prompt_next_item_or_confirm(update, context)
		await reply_text_with_hint(update.message, "Không xác định được món hiện tại. Gõ /start để đặt lại.")
		return ConversationHandler.END

	context.user_data["current_quantity"] = int(text)
	add_current_item_to_cart(context, context.user_data.get("current_toppings", []))
	return await prompt_next_item_or_confirm(update, context)


async def on_add_topping(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
	if not update.message:
		return ADD_TOPPING

	text = (update.message.text or "").strip()
	if wants_menu_image(text):
		await send_menu_image(update.message.chat)
		return ADD_TOPPING

	try:
		topping_ids = parse_topping_input(text)
	except ValueError as exc:
		await reply_text_with_hint(update.message, f"{exc}\n{ask_groq_for_order_prompt('topping_followup')}")
		return ADD_TOPPING

	pending_variants: List[tuple[str, int]] = context.user_data.pop("pending_size_variants", [])
	if pending_variants:
		current_item_id = context.user_data.get("current_item_id")
		if not current_item_id or current_item_id not in ITEM_BY_ID:
			await reply_text_with_hint(update.message, "Không xác định được món hiện tại. Gõ /start để đặt lại.")
			return ConversationHandler.END

		for size, qty in pending_variants:
			context.user_data["current_item_id"] = current_item_id
			context.user_data["current_size"] = size
			context.user_data["current_quantity"] = qty
			context.user_data["current_toppings"] = topping_ids
			add_current_item_to_cart(context, topping_ids)

		context.user_data["awaiting_topping_selection"] = False
		return await prompt_next_item_or_confirm(update, context)

	context.user_data["current_toppings"] = topping_ids
	context.user_data["awaiting_topping_selection"] = False
	if not topping_ids:
		# Nếu không chọn topping thì bỏ qua bước nhập số lượng, mặc định 1 ly.
		current_qty = context.user_data.get("current_quantity")
		if not current_qty:
			context.user_data["current_quantity"] = 1
		add_current_item_to_cart(context, [])
		return await prompt_next_item_or_confirm(update, context)

	await reply_text_with_hint(update.message, "Anh/chị muốn đặt topping này với số lượng bao nhiêu ạ?")
	return SELECT_QUANTITY


async def on_confirm_by_text(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
	if not update.message:
		return CONFIRM

	if wants_menu_image(update.message.text or ""):
		await send_menu_image(update.message.chat)
		return CONFIRM

	text = normalize_text(update.message.text or "")
	if text in {"huy", "huy don", "cancel"}:
		context.user_data.clear()
		await reply_text_with_hint(update.message, "Đã hủy đơn hàng.")
		return ConversationHandler.END

	if text in {"xac nhan", "xac nhan dat", "ok", "dong y"}:
		context.user_data["awaiting_payment_phone"] = True
		await reply_text_with_hint(update.message, 
			"Nhập số điện thoại thanh toán để tạo QR PayOS. "
			"Nội dung thanh toán sẽ là số điện thoại đó."
		)
		return PAYMENT_PHONE

	raw_text = (update.message.text or "").strip()

	# Fallback: nếu lệch state nhưng đang chờ số lượng topping ở bước CONFIRM thì vẫn xử lý số lượng.
	if context.user_data.get("awaiting_confirm_topping_quantity"):
		match = re.search(r"\b\d+\b", normalize_text(raw_text))
		if not match:
			await reply_text_with_hint(update.message, "Số lượng topping không hợp lệ. Vui lòng nhập số nguyên dương.")
			return CONFIRM

		quantity = int(match.group(0))
		if quantity <= 0:
			await reply_text_with_hint(update.message, "Số lượng topping không hợp lệ. Vui lòng nhập số nguyên dương.")
			return CONFIRM

		cart: List[Dict[str, object]] = context.user_data.get("cart", [])
		confirm_topping_ids: List[str] = context.user_data.get("pending_confirm_topping_ids", [])
		if not cart or not confirm_topping_ids:
			context.user_data.pop("awaiting_confirm_topping_quantity", None)
			context.user_data.pop("pending_confirm_topping_ids", None)
			await reply_text_with_hint(update.message, "Không tìm thấy món để thêm topping. Vui lòng chọn món lại.")
			return CONFIRM

		last_row = cart[-1]
		existing_toppings = [str(t) for t in last_row.get("toppings", [])]
		for topping_id in confirm_topping_ids:
			existing_toppings.extend([topping_id] * quantity)
		last_row["toppings"] = existing_toppings

		context.user_data.pop("awaiting_confirm_topping_quantity", None)
		context.user_data.pop("pending_confirm_topping_ids", None)

		summary = build_order_summary(context)
		await reply_text_with_hint(update.message, 
			f"Đã thêm topping cho món gần nhất.\n\n{summary}\n\n"
			"Bạn có muốn đặt thêm món không?\n"
			"- Nhập tên/mã/số món để thêm món\n"
			"- Nhập 'xác nhận' để sang bước thanh toán\n"
			"- Nhập 'hủy' để hủy đơn"
		)
		return CONFIRM

	if context.user_data.get("awaiting_confirm_topping_name"):
		try:
			confirm_topping_ids = parse_topping_input(raw_text)
		except ValueError as exc:
			await reply_text_with_hint(update.message, f"{exc}\nVui lòng nhập tên topping cần thêm hoặc gõ 'không'.")
			return CONFIRM

		if not confirm_topping_ids:
			context.user_data.pop("awaiting_confirm_topping_name", None)
			await reply_text_with_hint(update.message, "Đã bỏ qua thêm topping. Bạn có thể nhập món mới hoặc 'xác nhận'.")
			return CONFIRM

		context.user_data.pop("awaiting_confirm_topping_name", None)
		context.user_data["awaiting_confirm_topping_quantity"] = True
		context.user_data["pending_confirm_topping_ids"] = confirm_topping_ids
		await reply_text_with_hint(update.message, 
			f"Đã nhận topping: {format_topping_ids_for_text(confirm_topping_ids)}. "
			"Bạn muốn thêm topping này bao nhiêu phần? Vui lòng nhập số lượng."
		)
		return SELECT_QUANTITY

	if wants_add_topping_request(raw_text):
		cart: List[Dict[str, object]] = context.user_data.setdefault("cart", [])
		if not cart:
			await reply_text_with_hint(update.message, "Giỏ hàng đang trống, vui lòng nhập món trước khi thêm topping.")
			return CONFIRM

		context.user_data["awaiting_confirm_topping_name"] = True
		await reply_text_with_hint(update.message, "Bạn muốn thêm topping gì cho món gần nhất? (Gõ 'không' nếu không thêm nữa)")
		return CONFIRM

	# Ở bước CONFIRM, nếu user nhập topping thì hỏi số lượng topping.
	try:
		confirm_topping_ids = parse_topping_input(raw_text)
	except ValueError:
		confirm_topping_ids = []

	if confirm_topping_ids:
		cart: List[Dict[str, object]] = context.user_data.setdefault("cart", [])
		if not cart:
			await reply_text_with_hint(update.message, "Giỏ hàng đang trống, vui lòng nhập món trước khi thêm topping.")
			return CONFIRM

		context.user_data["awaiting_confirm_topping_quantity"] = True
		context.user_data["pending_confirm_topping_ids"] = confirm_topping_ids
		await reply_text_with_hint(update.message, 
			f"Đã nhận topping: {format_topping_ids_for_text(confirm_topping_ids)}. "
			"Bạn muốn thêm topping này bao nhiêu phần? Vui lòng nhập số lượng."
		)
		return SELECT_QUANTITY

	# Ở bước CONFIRM, cho phép user nhập thêm món để tiếp tục đặt hàng.
	try:
		new_item_ids = parse_item_selection_input(raw_text)
	except ValueError:
		await reply_text_with_hint(update.message, "Vui lòng nhập tên/mã/số món để thêm, hoặc nhập 'xác nhận' / 'hủy'.")
		return CONFIRM

	pending_item_ids: List[str] = context.user_data.setdefault("pending_item_ids", [])
	for item_id in new_item_ids:
		if item_id not in pending_item_ids:
			pending_item_ids.append(item_id)

	context.user_data["selected_item_ids"] = dedupe_preserve_order(
		context.user_data.get("selected_item_ids", []) + new_item_ids
	)
	context.user_data["awaiting_item_selection"] = False
	return await prompt_next_item_or_confirm(update, context)


async def on_payment_phone_input(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
	if not update.message:
		return PAYMENT_PHONE

	text = (update.message.text or "").strip()
	if wants_menu_image(text):
		await send_menu_image(update.message.chat)
		return PAYMENT_PHONE

	phone_number = sanitize_phone_number(text)
	if not phone_number:
		await reply_text_with_hint(update.message, "Số điện thoại không hợp lệ. Vui lòng nhập lại số điện thoại thanh toán.")
		return PAYMENT_PHONE

	await reply_text_with_hint(update.message, "Đang tạo QR PayOS, vui lòng chờ...")

	try:
		payment_data = await to_thread(create_payos_payment_link, context, phone_number)
	except Exception as exc:
		logger.exception("Loi tao QR PayOS: %s", exc)
		await reply_text_with_hint(update.message, f"Không tạo được QR PayOS: {exc}")
		return PAYMENT_PHONE

	qr_data = str(payment_data.get("qrCode") or payment_data.get("checkoutUrl") or "")
	checkout_url = str(payment_data.get("checkoutUrl") or "")
	payment_link_id = str(payment_data.get("paymentLinkId") or payment_data.get("id") or "")
	order_code = payment_data.get("orderCode")

	caption_lines = ["QR thanh toán PayOS đã tạo thành công."]
	if order_code is not None:
		caption_lines.append(f"Mã đơn: {order_code}")
	caption_lines.append(f"Nội dung: {phone_number}")
	if payment_link_id:
		caption_lines.append(f"Payment Link ID: {payment_link_id}")
	if checkout_url:
		caption_lines.append(f"Link thanh toán: {checkout_url}")

	if qr_data:
		await reply_photo_with_hint(update.message, photo=build_qr_image_url(qr_data), caption="\n".join(caption_lines))
	else:
		await reply_text_with_hint(update.message, "\n".join(caption_lines))

	if order_code is not None and update.effective_chat:
		asyncio.create_task(
			monitor_payos_payment(
				context=context,
				chat_id=update.effective_chat.id,
				order_code=str(order_code),
			)
		)

	context.user_data.clear()
	return ConversationHandler.END


async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
	context.user_data.clear()
	if update.message:
		await reply_text_with_hint(update.message, "Đã hủy đơn hàng.")
	return ConversationHandler.END


async def on_error(update: object, context: ContextTypes.DEFAULT_TYPE) -> None:
	err = context.error
	if isinstance(err, NetworkError):
		global _last_network_error_log_ts
		now = time.time()
		if now - _last_network_error_log_ts >= NETWORK_ERROR_LOG_COOLDOWN_SECONDS:
			logger.warning(
				"Loi mang tam thoi khi xu ly update: %s. Bot se tu thu lai ket noi.",
				err,
			)
			_last_network_error_log_ts = now
		return
	if isinstance(err, Conflict):
		logger.warning("Phat hien nhieu bot instance dang chay cung luc: %s", err)
		return
	logger.exception("Loi khong mong doi khi xu ly update", exc_info=err)


async def suggest_with_groq(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
	if not update.message:
		return

	request_text = " ".join(context.args).strip() if context.args else ""
	if not request_text:
		await reply_text_with_hint(update.message, 
			"Dùng: /goiy <tất cả yêu cầu của bạn>\n"
			"Ví dụ: /goiy tôi muốn món ít ngọt, mát lạnh, dưới 45k, thêm topping thạch"
		)
		return

	await reply_text_with_hint(update.message, "Đang gợi ý bằng Groq...")
	try:
		result = await to_thread(ask_groq_for_recommendation, request_text)
		await reply_text_with_hint(update.message, result)
	except Exception as exc:
		logger.exception("Loi goi Groq API: %s", exc)
		await reply_text_with_hint(update.message, "Không thể gọi Groq API lúc này. Bạn thử lại sau.")


async def suggest_from_text(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
	if not update.message:
		return

	text = (update.message.text or "").strip()
	if wants_menu_image(text):
		await send_menu_image(update.message.chat)
		if not is_order_in_progress(context):
			context.user_data["awaiting_item_selection"] = True
		return

	if context.user_data.get("awaiting_item_selection"):
		await on_select_items_by_text(update, context)
		return

	if context.user_data.get("awaiting_size_selection"):
		await on_select_size_by_text(update, context)
		return

	if context.user_data.get("awaiting_topping_selection"):
		await on_add_topping(update, context)
		return

	request_text = text
	if not request_text:
		return

	if is_order_in_progress(context):
		await reply_text_with_hint(update.message, 
			"Bạn đang trong luồng đặt hàng. Nếu muốn đổi món, bấm /cancel rồi /start.\n"
			"Nếu muốn AI gợi ý nhanh, dùng: /goiy <tất cả yêu cầu của bạn>"
		)
		return

	await reply_text_with_hint(update.message, "Đã nhận yêu cầu. Đang gợi ý bằng Groq...")
	try:
		result = await to_thread(ask_groq_for_recommendation, request_text)
		await reply_text_with_hint(update.message, result)
	except Exception as exc:
		logger.exception("Loi goi Groq API tu text thuong: %s", exc)
		await reply_text_with_hint(update.message, "Không thể gọi Groq API lúc này. Bạn thử lại sau.")


def build_app(token: str) -> Application:
	app = Application.builder().token(token).build()

	conv_handler = ConversationHandler(
		entry_points=[
			CommandHandler("start", start),
			CommandHandler("menu", show_main_menu),
		],
		states={
			SELECT_CATEGORY: [MessageHandler(filters.TEXT & ~filters.COMMAND, on_select_items_by_text)],
			SELECT_SIZE: [MessageHandler(filters.TEXT & ~filters.COMMAND, on_select_size_by_text)],
			SELECT_QUANTITY: [MessageHandler(filters.TEXT & ~filters.COMMAND, on_quantity_input)],
			ADD_TOPPING: [MessageHandler(filters.TEXT & ~filters.COMMAND, on_add_topping)],
			CONFIRM: [MessageHandler(filters.TEXT & ~filters.COMMAND, on_confirm_by_text)],
			PAYMENT_PHONE: [MessageHandler(filters.TEXT & ~filters.COMMAND, on_payment_phone_input)],
		},
		fallbacks=[CommandHandler("cancel", cancel)],
		allow_reentry=True,
	)

	app.add_handler(conv_handler)
	app.add_handler(CommandHandler("cancel", cancel))
	app.add_handler(CommandHandler("menu", show_main_menu))
	app.add_handler(CommandHandler("goiy", suggest_with_groq))
	app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, suggest_from_text))
	app.add_error_handler(on_error)
	return app


def get_bot_token() -> Optional[str]:
	token = os.getenv("TELEGRAM_BOT_TOKEN", "").strip()
	return token or None


def main() -> None:
	# Tạo event loop riêng để tránh lỗi đóng loop trên một số môi trường Python mới.
	token = get_bot_token()
	if not token:
		raise RuntimeError("Chưa có TELEGRAM_BOT_TOKEN trong môi trường.")

	app = build_app(token)
	loop = asyncio.new_event_loop()
	asyncio.set_event_loop(loop)
	try:
		loop.run_until_complete(app.run_polling(close_loop=False))
	finally:
		pending = asyncio.all_tasks(loop)
		for task in pending:
			task.cancel()
		if pending:
			loop.run_until_complete(asyncio.gather(*pending, return_exceptions=True))
		loop.close()

if __name__ == "__main__":
	main()

